
import os
import sys
import glob
import numpy as np
import pandas as pd
import yfinance as yf
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# RRG ENGINE (v1 Re-implementation)
# ─────────────────────────────────────────────
def calculate_rrg_coordinates(sector_series: pd.Series, bench_series: pd.Series) -> tuple:
    aligned_df = pd.DataFrame({'Sector': sector_series, 'Bench': bench_series}).dropna()
    if len(aligned_df) < 100:
        return pd.Series(100.0, index=sector_series.index), pd.Series(100.0, index=sector_series.index)
    
    rs = (aligned_df['Sector'] / aligned_df['Bench']) * 100
    ema_rs = rs.ewm(span=60, adjust=False).mean()
    std_rs = rs.rolling(window=60, min_periods=30).std()
    rs_ratio = 100 + 10 * ((rs - ema_rs) / std_rs.replace(0, np.nan)).fillna(0)
    
    d_ratio = rs_ratio.diff().fillna(0)
    ema_d_fast = d_ratio.ewm(span=10, adjust=False).mean()
    ema_d_slow = d_ratio.ewm(span=60, adjust=False).mean()
    std_d = d_ratio.rolling(window=60, min_periods=30).std()
    rs_momentum = 100 + 10 * ((ema_d_fast - ema_d_slow) / std_d.replace(0, np.nan)).fillna(0)
    
    return rs_ratio.reindex(sector_series.index).fillna(100.0), rs_momentum.reindex(sector_series.index).fillna(100.0)

def calculate_velocity_and_heading(ratio_s, mom_s):
    if len(ratio_s) < 2: return 0.0, 0.0
    dx = ratio_s.iloc[-1] - ratio_s.iloc[-2]
    dy = mom_s.iloc[-1] - mom_s.iloc[-2]
    vel = np.sqrt(dx**2 + dy**2)
    heading = (np.degrees(np.arctan2(dy, dx)) + 360) % 360
    return vel, heading

def get_sector_rankings(snap: pd.DataFrame, snapshot_path: str) -> pd.DataFrame:
    # Prefer data_all.csv if using snapshot_all.csv, else prefer data.csv
    if snapshot_path.endswith("snapshot_all.csv"):
        history_file = "data_all.csv" if os.path.exists("data_all.csv") else "data.csv"
    else:
        history_file = "data.csv" if os.path.exists("data.csv") else "data_all.csv"

    if not os.path.exists(history_file):
        raise FileNotFoundError(f"Historical data file ({history_file}) not found. Run analyzer first.")
    
    print(f"   Loading history from {history_file}...")
    hist = pd.read_csv(history_file)
    hist['datetime'] = pd.to_datetime(hist['datetime'])
    hist['symbols'] = hist['symbols'].str.upper().str.strip()
    
    # Get Benchmark
    print("   Fetching Nifty 50 benchmark...")
    bench = yf.download("^NSEI", period="1y", progress=False)['Close']
    if isinstance(bench, pd.DataFrame): bench = bench.squeeze()
    bench.index = pd.to_datetime(bench.index).tz_localize(None)

    # Create sector mapping and filter history
    sector_map = snap.set_index('symb')['sect'].to_dict()
    hist['sect'] = hist['symbols'].map(sector_map)
    
    # Synthesize Sector Indices (Equal Weighted)
    print("   Synthesizing sector indices...")
    sector_prices = hist.groupby(['datetime', 'sect'])['close'].mean().unstack().ffill()
    sector_prices.index = pd.to_datetime(sector_prices.index).tz_localize(None)

    sector_summary = []
    for sector in sector_prices.columns:
        if pd.isna(sector) or sector == 'Unknown': continue
        
        ratio_s, mom_s = calculate_rrg_coordinates(sector_prices[sector], bench)
        
        r = ratio_s.iloc[-1]
        m = mom_s.iloc[-1]
        vel, head = calculate_velocity_and_heading(ratio_s, mom_s)
        
        # ── Aggregating Breadth Metrics from Snapshot ──
        sec_snap = snap[snap['sect'] == sector]
        if sec_snap.empty: continue

        # 1. RSI Breadth (User Suggestion: % of stocks with RSI >= 50)
        rsi_breadth = (sec_snap['rsi'] >= 50).mean() * 100
        
        # 2. EMA Breadth (% stocks above SMA50)
        ema_breadth = (sec_snap['g050'] == True).mean() * 100
        
        # 3. Delivery Conviction (Average delivery %)
        avg_delivery = sec_snap['DlPer'].mean()

        # 4. Heading Direction Score (Rewards NE movement towards Leading quadrant)
        # Formula: cos(heading - 45 degrees) mapped to 0-100
        heading_score = max(0, np.cos(np.radians(head - 45))) * 100

        # ── Weighted Multi-Factor Rotational Score (Institutional Grade) ──
        # Weights: RSI Breadth(25%), EMA Breadth(30%), Heading(25%), Delivery(20%)
        score = (rsi_breadth * 0.25) + (ema_breadth * 0.30) + (heading_score * 0.25) + (avg_delivery * 0.20)

        # Quadrant Determination
        quad = "LEADING" if r >= 100 and m >= 100 else "WEAKENING" if r >= 100 else "LAGGING" if m < 100 else "IMPROVING"
        
        sector_summary.append({
            'Sector Name': sector,
            'RRG Quadrant': quad,
            'RS-Ratio': round(r, 2),
            'RS-Momentum': round(m, 2),
            'Rotational Score': round(score, 2),
            'Velocity': round(vel, 2),
            'Heading': round(head, 1)
        })

    df = pd.DataFrame(sector_summary).sort_values("Rotational Score", ascending=False)
    print(f"   Sectors  : {len(df)} calculated manually")
    return df

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
DARKSOUL_XLSX     = "sector_rotation_multi_report.xlsx"
OUTPUT_FILE       = "swing_scanner_report.xlsx"

# Sector filter — which quadrants to include
GOOD_QUADRANTS    = {"LEADING", "IMPROVING"}

# Stock filters
MIN_RSI           = 45
MAX_RSI           = 72
MIN_ADX           = 20
MIN_RVOL          = 1.0

BULLISH_PATTERNS  = {
    "Inverse Head and Shoulders", "Double Bottom", "Triple Bottom",
    "Ascending Triangle", "Flag Pattern (Bull Flag)", "Pennant Pattern",
    "Falling Wedge", "Cup and Handle", "Rounding Bottom",
    "Long-term Cup and Handle", "Long-term Rounding Bottom",
    "Rectangle Pattern", "Symmetrical Triangle",
}
BEARISH_PATTERNS  = {
    "Head and Shoulders", "Double Top", "Triple Top",
    "Descending Triangle", "Flag Pattern (Bear Flag)",
    "Rising Wedge", "Rounding Top", "Diamond Top",
}

# ─────────────────────────────────────────────
# STEP 1 — Load snapshot.csv
# ─────────────────────────────────────────────
def load_snapshot(file_path: str) -> pd.DataFrame:
    print(f"   Snapshot : {file_path}")
    df = pd.read_csv(file_path).fillna(np.nan)
    df.columns = df.columns.str.strip()
    if 'symb' in df.columns:
        df['symb'] = df['symb'].astype(str).str.upper().str.strip()
    return df

# ─────────────────────────────────────────────
# STEP 3 — Filter + merge
# ─────────────────────────────────────────────
def build_candidates(snap: pd.DataFrame, sectors: pd.DataFrame) -> tuple:

    # Leading / Improving sectors only
    good_sectors = sectors[sectors['RRG Quadrant'].isin(GOOD_QUADRANTS)]['Sector Name'].tolist()
    print(f"\n   ✅ Good sectors ({len(good_sectors)}): {good_sectors}")
    steps = {}

    if not good_sectors:
        print("   ⚠️  No Leading/Improving sectors today — no trades.")
        return pd.DataFrame(), steps

    df = snap.copy()
    date_col = 'date' if 'date' in df.columns else None
    symb_col = 'symb' if 'symb' in df.columns else None

    # ── Sector filter ──
    sect_col = 'sect' if 'sect' in df.columns else None
    if sect_col:
        df = df[df[sect_col].isin(good_sectors)]
        steps["Filter - Sector"] = df[[c for c in [symb_col, sect_col] if c]].copy()
    print(f"   After sector filter    : {len(df)}")

    # ── Stage 2 only ──
    if 'stge' in df.columns:
        df = df[df['stge'].astype(str).str.contains("Stage 2", na=False)]
        steps["Filter - Stage 2"] = df[[c for c in [symb_col, sect_col, 'stge'] if c]].copy()
    print(f"   After Stage 2 filter   : {len(df)}")

    # ── Above SMA200 ──
    if 'g200' in df.columns:
        df = df[pd.to_numeric(df['g200'], errors='coerce').fillna(-999) > 0]
        steps["Filter - SMA200"] = df[[c for c in [symb_col, sect_col, 'g200'] if c]].copy()
    print(f"   After SMA200 filter    : {len(df)}")

    # ── RSI range ──
    if 'rsi' in df.columns:
        rsi = pd.to_numeric(df['rsi'], errors='coerce')
        df  = df[(rsi >= MIN_RSI) & (rsi <= MAX_RSI)]
        steps["Filter - RSI"] = df[[c for c in [symb_col, sect_col, 'rsi'] if c]].copy()
    print(f"   After RSI filter       : {len(df)}")

    # ── ADX ──
    if 'adx' in df.columns:
        df = df[pd.to_numeric(df['adx'], errors='coerce').fillna(0) >= MIN_ADX]
        steps["Filter - ADX"] = df[[c for c in [symb_col, sect_col, 'adx'] if c]].copy()
    print(f"   After ADX filter       : {len(df)}")

    # ── Relative Volume ──
    if 'rvol' in df.columns:
        df = df[pd.to_numeric(df['rvol'], errors='coerce').fillna(0) >= MIN_RVOL]
        steps["Filter - RVOL"] = df[[c for c in [symb_col, sect_col, 'rvol'] if c]].copy()
    print(f"   After RVOL filter      : {len(df)}")

    # ── Pattern filter ──
    pat_col = next((c for c in ['mpat', 'patt'] if c in df.columns), None)
    if pat_col:
        df = df[~df[pat_col].isin(BEARISH_PATTERNS)]
        steps["Filter - Pattern"] = df[[c for c in [symb_col, sect_col, pat_col] if c]].copy()
        df = df.copy()
        df['pat_bullish'] = df[pat_col].isin(BULLISH_PATTERNS).astype(int)
    else:
        df['pat_bullish'] = 0
    print(f"   After pattern filter   : {len(df)}")

    if df.empty:
        return df, steps

    # ── Merge sector RRG data ──
    sector_info = sectors[['Sector Name', 'RRG Quadrant', 'RS-Ratio',
                            'RS-Momentum', 'Rotational Score', 'Velocity', 'Heading']].copy()
    sector_info = sector_info.rename(columns={'Sector Name': sect_col})
    df = df.merge(sector_info, on=sect_col, how='left')

    # ── Sort: pat_bullish → rotational score → rvol ──
    sort_cols = [c for c in ['pat_bullish', 'Rotational Score', 'rvol'] if c in df.columns]
    df = df.sort_values(sort_cols, ascending=False)

    # ── Final output columns ──
    keep = []
    for col in ['date', 'symb', sect_col, 'RRG Quadrant', 'Rotational Score', 'RS-Ratio', 'RS-Momentum',
                'clos', 'rsi', 'adx', 'stge', 'rvol', 'DlPer',
                'g200', 'g050', 'zone', 'obv',
                'mpat', 'pcon', 'patt', 'pat_bullish',
                'bbsq', 'bbbw', 'mcdl', 'vola', 'delt',
                'shgh', 'slw', 'tren', 'tstr']:
        if col and col in df.columns:
            keep.append(col)
    return df[keep], steps

# ─────────────────────────────────────────────
# STEP 4 — Styling Layer
# ─────────────────────────────────────────────
def apply_corporate_styling(file_path):
    wb = openpyxl.load_workbook(file_path)
    cell_font = Font(name="Segoe UI", size=10)
    
    for i, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        
        # ── Convert to Table ──
        if ws.max_row > 1:
            last_col = get_column_letter(ws.max_column)
            tab_range = f"A1:{last_col}{ws.max_row}"
            # Display name must be unique and contain no spaces
            safe_name = "".join(filter(str.isalnum, sheet_name)) + f"_{i}"
            tab = Table(displayName=f"Table_{safe_name}", ref=tab_range)
            style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

        # Column width auto-fit
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
                cell.font = cell_font
            ws.column_dimensions[column].width = min(max_length + 3, 40)
            
    wb.save(file_path)

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    print("\n" + "="*60)
    print("🚀 Swing Scanner Bridge v3.0 (Excel Output)")
    print("="*60)

    # ── resolve CSV path ──
    if len(sys.argv) > 1:
        snapshot_path = sys.argv[1]
    else:
        # Sort files by modification time so the most recently created file is truly last
        snapshot_files = sorted([f for f in os.listdir(".") if f.endswith("snapshot.csv") and not f.endswith("snapshot_all.csv")], key=os.path.getmtime)
        snapshot_all_files = sorted([f for f in os.listdir(".") if f.endswith("snapshot_all.csv")], key=os.path.getmtime)

        if snapshot_files and snapshot_all_files:
            print("Choose snapshot source:")
            print("1. snapshot.csv")
            print("2. snapshot_all.csv")
            choice = input("Enter choice [default 1]: ").strip()
            if choice == '2':
                snapshot_path = snapshot_all_files[-1]
            else:
                snapshot_path = snapshot_files[-1]
        elif snapshot_files:
            snapshot_path = snapshot_files[-1]
        elif snapshot_all_files:
            snapshot_path = snapshot_all_files[-1]
        else:
            print("❌ No snapshot CSV found in current directory.")
            sys.exit(1)

    print(f"\n📂 Loading files from {snapshot_path}...")
    try:
        snap    = load_snapshot(snapshot_path)
        sectors = get_sector_rankings(snap, snapshot_path)
    except Exception as e:
        print(f"❌ Error during manual RRG calculation: {e}")
        sys.exit(1)

    print("\n📊 Sector RRG snapshot (darksoul):")
    print(sectors[['Sector Name', 'RRG Quadrant', 'RS-Ratio',
                   'RS-Momentum', 'Rotational Score']].to_string(index=False))
    print("\n🔍 Filtering swing candidates...")
    candidates, filter_steps = build_candidates(snap, sectors)

    # Determine output filename based on snapshot type
    suffix = "_all" if snapshot_path.endswith("snapshot_all.csv") else ""
    final_output_file = OUTPUT_FILE.replace(".xlsx", f"{suffix}.xlsx")

    # Write to Excel
    print(f"\n💾 Compiling Master Report: {final_output_file}")
    with pd.ExcelWriter(final_output_file, engine='openpyxl') as writer:
        candidates.to_excel(writer, sheet_name="Swing Candidates", index=False)
        sectors.to_excel(writer, sheet_name="Sector RRG Snapshot", index=False)
        # Add bifurcated filter sheets
        for sheet_name, step_df in filter_steps.items():
            step_df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    if candidates.empty:
        print("\n⚠️  No final candidates today, but filter progression saved to Excel.")
    else:
        print(f"✅ {len(candidates)} candidates found.")

    apply_corporate_styling(final_output_file)

    today = datetime.now().strftime("%d-%b-%Y")
    print(f"✅ {len(candidates)} candidates saved and styled.")
    print(f"\n🏆 Top 10 — {today}:")
    print(candidates[['symb', sect_col if (sect_col := 'sect') else 'sect',
                       'RRG Quadrant', 'Rotational Score',
                       'clos', 'rsi', 'adx', 'mpat']].head(10).to_string(index=False))
    print("\n" + "="*60)
     
if __name__ == "__main__":
    main()
