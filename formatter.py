import pandas as pd
import os
import sys
import glob
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- COLOR CONSTANTS ---
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
BLUE_FILL  = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
DARK_GREEN_FILL = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
DARK_RED_FILL = PatternFill(start_color='9C0006', end_color='9C0006', fill_type='solid')
NEUTRAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
WHITE_FONT = Font(color='FFFFFF')
YELLOW_FILL = PatternFill(start_color='FFEB84', end_color='FFEB84', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FCAA67', end_color='FCAA67', fill_type='solid')
F_BLUE = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
def colorize_snapshot(csv_path):
    if not os.path.exists(csv_path):
        print(f"Error: File {csv_path} not found.")
        return

    xlsx_path = csv_path.replace('.csv', '.xlsx')
    print(f"Processing {csv_path}...")

    # Load Data
    df = pd.read_csv(csv_path)
    
    # Split Patterns and SOHLC to separate sheets
    pattern_cols = ['psta', 'pend', 'ppnt', 'patt']
    sohlc_cols = ['open', 'shgh', 'slw', 'high', 'low', 'eqb', 's020', 's050', 's100', 's200', 'h52h', 'l52l']
    
    # Check if cols exist before splitting
    actual_pattern_cols = [c for c in pattern_cols if c in df.columns]
    actual_sohlc_cols = [c for c in sohlc_cols if c in df.columns]
    
    df_main = df.drop(columns=actual_pattern_cols + actual_sohlc_cols)
    df_patterns = df[['symb'] + actual_pattern_cols].copy() if actual_pattern_cols else pd.DataFrame()
    df_sohlc = df[['symb'] + actual_sohlc_cols].copy() if actual_sohlc_cols else pd.DataFrame()

    # Create Excel Writer
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df_main.to_excel(writer, sheet_name='Main Analysis', index=False)
        if not df_patterns.empty:
            df_patterns.to_excel(writer, sheet_name='Chart Patterns', index=False)
        if not df_sohlc.empty:
            df_sohlc.to_excel(writer, sheet_name='SOHLC', index=False)
        
        # Legend Sheet Data
        legend_data = [
            ["Column", "Condition", "Color", "Meaning"],
            ["Booleans", "True", "Light Green", "Bullish / Signal Active"],
            ["Booleans", "False", "Light Red", "Bearish / Signal Inactive"],
            ["chan", "> 0", "Light Green", "Price Gain"],
            ["CMF_20", "> 0.05", "Light Green", "Money Inflow"],
            ["SUPERTd", "1.0", "Light Green", "Uptrend Confirm"],
            ["SQZ_ON", "1 / TRUE", "Light Green", "Coiling / Squeeze"],
            ["tren", "Uptrend", "Light Green", "Bullish Trend"],
            ["rsi", "Intensity Gradient", "G -> Y -> R", "Momentum strength vs Overbought risk"],
            ["vola", "Inverse Gradient", "G -> Y -> R", "Stability vs Annualized Volatility Risk"],
            ["bbbw", "Intensity Gradient", "G -> Y -> R", "Tight Squeeze to Extreme Expansion"],
            ["STOCHk", "Intensity Gradient", "G -> Y -> R", "Oversold Accumulation to Overbought Pivot"],
            ["adx", ">= 25", "Light Green", "Strong Trend"],
            ["rvol", ">= 3.0", "Dark Green", "Surge: 3x normal volume"],
            ["vrnk/arnk", "<= 50", "Dark Green", "Market Liquidity Leader"],
            ["delt", ">= 30%", "Dark Green / White Font", "Significant room to grow"],
            ["delt", "<= 7%", "Light Red", "Near 52W High (Mean Reversion Risk)"],
            ["zone", "Premium", "Dark Red", "Overbought Highs (Mean Reversion Risk)"],
            ["zone", "Near Discount", "Dark Green", "Accumulation Zone (Near Support)"],
            ["zone", "Equilibrium", "Yellow", "Mid-range Consolidation"],
            ["MT_Zone", "Percentile Position", "Same as zone", "Position within the 60-day swing structure"]
        ]
        pd.DataFrame(legend_data).to_excel(writer, sheet_name='Legend', index=False, header=False)

    # Re-open with openpyxl for styling
    wb = load_workbook(xlsx_path)
    ws = wb['Main Analysis']

    # 1. Freeze Top Row
    ws.freeze_panes = 'A2'

    # 2. Setup References
    headers = [cell.value for cell in ws[1]]
    def get_col_ref(name):
        try:
            idx = headers.index(name) + 1
            letter = get_column_letter(idx)
            return f"{letter}2:{letter}{ws.max_row}"
        except ValueError: return None

    def get_col_let(name):
        try:
            return get_column_letter(headers.index(name) + 1)
        except ValueError: return None

    # 3. Conditional Formatting Rules (Professional Engine)
    # Red-Yellow-Green Scales (High is Good)
    for col in ["chan", "rvol", "ascr", "adx", "score", "DlPer", "STOCHk_14_3_3", "EFI_13", "pcon", "CMF_20", "volu", "obv"]:
        ref = get_col_ref(col)
        if ref:
            ws.conditional_formatting.add(ref, ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            ))

    # Inverse Scales (High is Bad/Risk)
    for col in ["vola", "delt", "arnk", "vrnk", "rrnk", "bbbw"]:
        ref = get_col_ref(col)
        if ref:
            ws.conditional_formatting.add(ref, ColorScaleRule(
                start_type='min', start_color='63BE7B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='F8696B'
            ))

    # Signal Highlighting (Booleans & Numeric Flags)
    for col in ["bbup", "bbdn", "bbsq", "vspk", "g200", "g050", "g020"]:
        ref = get_col_ref(col)
        if ref:
            # Match literal text "True" or "False" as requested
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"True"'], fill=GREEN_FILL))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"TRUE"'], fill=GREEN_FILL))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['TRUE'], fill=GREEN_FILL))
            
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"False"'], fill=RED_FILL))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['"FALSE"'], fill=RED_FILL))
            ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=['FALSE'], fill=RED_FILL))

    # Numeric flags remain on their own logic
    for col in ["SUPERTd_7_3.0", "SQZ_ON", "SQZ_OFF"]:
        ref = get_col_ref(col)
        if ref:
            ws.conditional_formatting.add(ref, CellIsRule(operator='greaterThan', formula=['0'], fill=GREEN_FILL))
            ws.conditional_formatting.add(ref, CellIsRule(operator='lessThanOrEqual', formula=['0'], fill=RED_FILL))

    # Categorical logic
    for col, mapping in [
        ("tren", {"Uptrend": GREEN_FILL, "Downtrend": RED_FILL, "Sideways": YELLOW_FILL}),
        ("tstr", {"Strong": GREEN_FILL, "Weak": RED_FILL, "Moderate": YELLOW_FILL}),
        ("vtrd", {"Increasing": GREEN_FILL, "Decreasing": RED_FILL}),
        ("zone", {
            "Premium": RED_FILL, 
            "Discount": GREEN_FILL,
            "Equilibrium": YELLOW_FILL,
            "Near Discount": DARK_GREEN_FILL,
            "Near Premium": DARK_RED_FILL
        }),
        ("MT_Zone", {
            "Premium": RED_FILL, 
            "Discount": GREEN_FILL,
            "Equilibrium": YELLOW_FILL,
            "Near Discount": DARK_GREEN_FILL,
            "Near Premium": DARK_RED_FILL
        }),
        ("stge", {
            "Stage 2 (Uptrend)": GREEN_FILL,
            "Stage 4 (Downtrend)": RED_FILL,
            "Stage 1/3 (Neutral)": YELLOW_FILL
        }),
        ("screener_type", {"Intraday": BLUE_FILL, "Swing": PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')})
    ]:
        ref = get_col_ref(col)
        if ref:
            for val, fill in mapping.items():
                # Apply white font for dark backgrounds
                font = Font(color='FFFFFF') if fill in [DARK_GREEN_FILL, DARK_RED_FILL] else None
                ws.conditional_formatting.add(ref, CellIsRule(operator='equal', formula=[f'"{val}"'], fill=fill, font=font))

    # wrsi: >80 Blue, 70-80 Dark Green, 60-70 Green, 50-60 Yellow, 40-50 Orange, <40 Red
    ref_wrsi = get_col_ref("wrsi")
    if ref_wrsi:
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='greaterThan', formula=['80'], fill=DARK_GREEN_FILL, font=Font(color='FFFFFF'))) # Using DARK_GREEN_FILL as proxy for blue if not defined
        f_blue = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='greaterThan', formula=['80'], fill=f_blue, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['70', '80'], fill=DARK_GREEN_FILL, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['60', '70'], fill=GREEN_FILL))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['50', '60'], fill=YELLOW_FILL))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='between', formula=['40', '50'], fill=ORANGE_FILL))
        ws.conditional_formatting.add(ref_wrsi, CellIsRule(operator='lessThan', formula=['40'], fill=RED_FILL))

    # rsi / RSI_2: >80 Red, 70-80 Orange, 40-70 Yellow, 20-40 Green, <20 Dark Green
    for col in ["rsi", "RSI_2"]:
        ref_rsi = get_col_ref(col)
        if ref_rsi:
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='greaterThan', formula=['80'], fill=RED_FILL))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='between', formula=['70', '80'], fill=ORANGE_FILL))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='between', formula=['40', '70'], fill=YELLOW_FILL))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='between', formula=['20', '40'], fill=GREEN_FILL))
            ws.conditional_formatting.add(ref_rsi, CellIsRule(operator='lessThan', formula=['20'], fill=DARK_GREEN_FILL, font=Font(color='FFFFFF')))

    # Williams %R Specific Mapping (6-Tier Momentum Scale)
    ref = get_col_ref("WILLR_14")
    if ref:
        f_extreme = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid') # Blue
        f_vstrong = PatternFill(start_color='375623', end_color='375623', fill_type='solid') # Dark Green
        f_bullish = PatternFill(start_color='63BE7B', end_color='63BE7B', fill_type='solid') # Green
        f_neutral = PatternFill(start_color='FFEB84', end_color='FFEB84', fill_type='solid') # Yellow
        f_weak    = PatternFill(start_color='FCAA67', end_color='FCAA67', fill_type='solid') # Orange
        f_red     = PatternFill(start_color='F8696B', end_color='F8696B', fill_type='solid') # Red

        ws.conditional_formatting.add(ref, CellIsRule(operator='greaterThan', formula=['-10'], fill=f_extreme, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-20', '-10'], fill=f_vstrong, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-50', '-20'], fill=f_bullish))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-80', '-50'], fill=f_neutral))
        ws.conditional_formatting.add(ref, CellIsRule(operator='between', formula=['-90', '-80'], fill=f_weak))
        ws.conditional_formatting.add(ref, CellIsRule(operator='lessThan', formula=['-90'], fill=f_red))

    # 4. Trend Proximity and Stochastic Cross Logic
    c_let = get_col_let("clos")
    
    # SUPERT_7_3.0 Trend Strength Mapping
    s_let, s_ref = get_col_let("SUPERT_7_3.0"), get_col_ref("SUPERT_7_3.0")
    if c_let and s_let and s_ref:
        dist = f"((${c_let}2-${s_let}2)/${s_let}2)*100"
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"{dist}>5"], fill=F_BLUE, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>2,{dist}<=5)"], fill=DARK_GREEN_FILL, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>0,{dist}<=2)"], fill=GREEN_FILL))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>=-2,{dist}<=0)"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"AND({dist}>=-5,{dist}<-2)"], fill=ORANGE_FILL))
        ws.conditional_formatting.add(s_ref, FormulaRule(formula=[f"{dist}<-5"], fill=RED_FILL))

    # EMA_21 Trend Strength Mapping
    e_let, e_ref = get_col_let("EMA_21"), get_col_ref("EMA_21")
    if c_let and e_let and e_ref:
        dist = f"((${c_let}2-${e_let}2)/${e_let}2)*100"
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"{dist}>8"], fill=F_BLUE, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>3,{dist}<=8)"], fill=DARK_GREEN_FILL, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>0,{dist}<=3)"], fill=GREEN_FILL))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>-3,{dist}<=0)"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"AND({dist}>=-8,{dist}<-3)"], fill=ORANGE_FILL))
        ws.conditional_formatting.add(e_ref, FormulaRule(formula=[f"{dist}<-8"], fill=RED_FILL))

    # ws30 Trend Strength Mapping
    w_let, w_ref = get_col_let("ws30"), get_col_ref("ws30")
    if c_let and w_let and w_ref:
        dist = f"((${c_let}2-${w_let}2)/${w_let}2)*100"
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"{dist}>15"], fill=F_BLUE, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>5,{dist}<=15)"], fill=DARK_GREEN_FILL, font=Font(color='FFFFFF')))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>0,{dist}<=5)"], fill=GREEN_FILL))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>=-5,{dist}<=0)"], fill=YELLOW_FILL))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"AND({dist}>=-15,{dist}<-5)"], fill=ORANGE_FILL))
        ws.conditional_formatting.add(w_ref, FormulaRule(formula=[f"{dist}<-15"], fill=RED_FILL))

    # Stochastic K vs D Crossing and Overbought/Oversold Logic
    k_let, k_ref = get_col_let("STOCHk_14_3_3"), get_col_ref("STOCHk_14_3_3")
    d_let, d_ref = get_col_let("STOCHd_14_3_3"), get_col_ref("STOCHd_14_3_3")
    if k_let and d_let and k_ref:
        # Apply to K and D columns if both exist
        for ref in [k_ref, d_ref] if d_ref else [k_ref]:
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"AND(${k_let}2>${d_let}2,${k_let}2>80)"], fill=F_BLUE, font=Font(color='FFFFFF')))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"AND(${k_let}2<${d_let}2,${k_let}2<20)"], fill=RED_FILL))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"${k_let}2>${d_let}2"], fill=GREEN_FILL))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"${k_let}2<${d_let}2"], fill=ORANGE_FILL))
            ws.conditional_formatting.add(ref, FormulaRule(formula=[f"ABS(${k_let}2-${d_let}2)<=2"], fill=YELLOW_FILL))

    # 5. Global Formatting Refinements
    for i, col in enumerate(ws.columns, 1):
        col_name = headers[i-1]
        # Bold headers
        ws.cell(row=1, column=i).font = Font(bold=True)
        # Number formatting
        if col_name == 'volu':
            for cell in col[1:]:
                cell.number_format = '#,##0'

    # Auto-fit Column Widths
    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(i)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 40) # Cap width

    # Style Legend Sheet
    if 'Legend' in wb.sheetnames:
        lws = wb['Legend']
        for row in lws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
        lws.column_dimensions['A'].width = 20
        lws.column_dimensions['B'].width = 30
        lws.column_dimensions['C'].width = 25
        lws.column_dimensions['D'].width = 40

    # 3. Convert to Excel Tables for Filtering
    for sheet_name, table_name in [('Main Analysis', 'AnalysisTable'), ('Chart Patterns', 'PatternsTable'), ('SOHLC', 'SOHLCTable')]:
        if sheet_name in wb.sheetnames:
            target_ws = wb[sheet_name]
            if target_ws.max_row > 1:
                # Define the table range (e.g., A1:Z100)
                last_col = get_column_letter(target_ws.max_column)
                tab_range = f"A1:{last_col}{target_ws.max_row}"
                
                tab = Table(displayName=table_name, ref=tab_range)
                # Use a light style so it doesn't clash with our conditional colors
                style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=False, showColumnStripes=False)
                tab.tableStyleInfo = style
                target_ws.add_table(tab)

    wb.save(xlsx_path)
    print(f"Done! Excel file saved: {xlsx_path}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        target = sys.argv[1]
    else:
        # Sort by OS modification time to ensure we actually get the most recently created file
        snapshot_files = sorted(glob.glob("*snapshot.csv"), key=os.path.getmtime)
        snapshot_all_files = sorted(glob.glob("*snapshot_all.csv"), key=os.path.getmtime)
        # Filter out 'all' snapshots from the standard list to avoid duplicates
        snapshot_files = [f for f in snapshot_files if not f.endswith("_all.csv")]

        if snapshot_files and snapshot_all_files:
            print("Choose snapshot source:")
            print("1. snapshot.csv")
            print("2. snapshot_all.csv")
            choice = input("Enter choice [default 1]: ").strip()
            if choice == '2':
                target = snapshot_all_files[-1]
            else:
                target = snapshot_files[-1]
        elif snapshot_files:
            target = snapshot_files[-1]
        elif snapshot_all_files:
            target = snapshot_all_files[-1]
        else:
            print("No snapshot CSV found in current directory.")
            sys.exit(1)
    colorize_snapshot(target)
